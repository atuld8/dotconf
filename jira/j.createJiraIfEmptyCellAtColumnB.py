import argparse
import os
import requests
import openpyxl

def get_jira_server_url():
    jira_server = os.environ.get("JIRA_SERVER_NAME")
    if not jira_server:
        raise EnvironmentError("JIRA_SERVER_NAME environment variable is not set.")
    return f"https://{jira_server}/rest/api/2"

def prompt_for_token():
    token = os.environ.get("JIRA_ACC_TOKEN")
    if not token:
        token = input("Enter your Jira Personal Access Token: ")
    return token

def create_jira_issue(summary, description, token):
    url = f"{get_jira_server_url()}/issue"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    # The caller will pass all required fields as a dict
    data = {
        "fields": summary  # summary here is actually the full fields dict
    }

    response = requests.post(url, headers=headers, json=data, timeout=10)
    if response.ok:
        print(f"Created Jira issue: {response.json()['key']}")
        return response.json()['key']
    else:
        print(f"Failed to create Jira issue: {response.text}")
        return None


def clean_value(value):
    """
    Cleans a given value by removing newline characters and leading/trailing whitespace.

    Args:
        value: The input value to clean. Can be any type.

    Returns:
        str: The cleaned string representation of the input value. Returns an empty string if the input is falsy.
    """ 
    if not value:
        return ""
    return str(value).replace('\n', '').strip()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("-f", "--file", required=True, help="Excel file path")
    parser.add_argument("-s", "--sheet", required=True, help="Sheet name")
    parser.add_argument("-d", "--dry-run", action="store_true", help="Print data only, do not create Jira issues")
    args = parser.parse_args()

    token = prompt_for_token()
    wb = openpyxl.load_workbook(args.file)
    ws = wb[args.sheet]

    # Read header row B2-T2 for required Jira properties
    header_props = [ws.cell(row=2, column=col_idx).value for col_idx in range(2, 21)]
    missing_headers = [f"Column {chr(64+col_idx)}2" for col_idx, val in zip(range(2, 21), header_props) if not val]
    if missing_headers:
        raise ValueError(f"Missing required Jira property in header: {', '.join(missing_headers)}")

    # Fetch Jira field metadata once and cache mapping
    def jira_request_for_field(token):
        url = f"{get_jira_server_url()}/field"
        headers = {"Authorization": f"Bearer {token}"}
        response = requests.get(url, headers=headers, timeout=10)
        return response.json() if response.ok else None

    fields_response = jira_request_for_field(token)
    display_name_to_id = {f["name"]: f["id"] for f in fields_response} if fields_response else {}

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        col_b = clean_value(row[1].value)  # B column
        col_c = clean_value(row[2].value)  # C column
        if not col_b and col_c:
            # Build fields dict from B2-T2 headers and row values
            fields = {}

            for idx, header in enumerate(header_props):
                if header:
                    field_id = display_name_to_id.get(header)
                    if field_id:
                        value = clean_value(row[idx+1].value)
                        fields[field_id] = value
            # Always set required Jira fields with literal keys
            fields["project"] = {"key": "ROSTER", "name": fields["project"]}
            fields["issuetype"] = {"name": fields["issuetype"]}
            fields["customfield_21611"] = {"name": fields["customfield_21611"]}
            fields["customfield_21609"] = {"name": fields["customfield_21609"]}
            fields["customfield_21607"] = {"name": fields["customfield_21607"]}
            fields["customfield_21605"] = {"name": fields["customfield_21605"]}
            fields["customfield_21603"] = {"name": fields["customfield_21603"]}
            fields["customfield_21601"] = {"name": fields["customfield_21601"]}
            fields["customfield_12671"] = {"value": fields["customfield_12671"]}

            if not args.dry_run:
                issue_key = create_jira_issue(fields, None, token)
                if issue_key:
                    row[1].value = issue_key  # Update B column with new issue key
            else:
                print(f"Dry run: Would create Jira issue with fields: \n{fields}")

    if not args.dry_run:
        wb.save(args.file)

if __name__ == "__main__":
    main()
