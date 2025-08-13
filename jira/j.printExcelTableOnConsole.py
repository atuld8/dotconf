import argparse
import openpyxl
import openpyxl.worksheet.formula
import sys

def clean_cell(val):
    if val is None:
        return ""
    return str(val).strip().replace('\n', ' ')

def print_table(headers, rows):
    # Remove leading/trailing newlines and spaces
    headers = [clean_cell(h) for h in headers]
    data_rows = [[clean_cell(cell) for cell in row] for row in rows]
    # Calculate column widths
    col_widths = [max(len(str(h)), *(len(str(row[i])) for row in data_rows)) for i, h in enumerate(headers)]
    # Line separator
    line_sep = '|' + '|'.join('-' * (w + 2) for w in col_widths) + '|'
    print(line_sep)
    # Print header
    header_line = '| ' + ' | '.join(h.ljust(col_widths[i]) for i, h in enumerate(headers)) + ' |'
    print(header_line)
    print(line_sep)
    # Print data rows
    for row in data_rows:
        print('| ' + ' | '.join(str(row[i]).ljust(col_widths[i]) for i in range(len(headers))) + ' |')
    print(line_sep)
    print(f"Rows: {len(data_rows)}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', required=True, help='Excel file path')
    parser.add_argument('-s', '--sheet', help='Sheet name')
    parser.add_argument('-t', '--header', default="A1", help='Header start cell, e.g. B2')
    parser.add_argument('-p', '--print-headers', help='Comma-separated list of headers to print (case-sensitive, matches Excel header values)')
    parser.add_argument('-o', '--only-headers', action='store_true', help='Print only headers in vertical order')
    parser.add_argument('-m', '--formula-mode', choices=['print', 'convert'], default='print', help='Choose to print formula as text or convert to value (default: print)')
    parser.add_argument('-e', '--skip-if-empty', help='Skip row if the value of cell is empty for this header column')
    args = parser.parse_args()

    try:
        wb = openpyxl.load_workbook(args.file)
    except Exception as e:
        print(f'Error loading workbook: {e}', file=sys.stderr)
        sys.exit(1)

    if not args.sheet:
        print("Sheets in Excel file:")
        for sheet in wb.sheetnames:
            print(f"- {sheet}")
        sys.exit(0)

    try:
        ws = wb[args.sheet]
    except Exception as e:
        print(f'Error loading sheet: {e}', file=sys.stderr)
        sys.exit(1)

    # Parse header start cell
    import re
    m = re.match(r'([A-Z]+)(\d+)', args.header.upper())
    if not m:
        print('Invalid header start cell format. Use e.g. B2.', file=sys.stderr)
        sys.exit(1)
    col_letter, row_num = m.groups()
    row_num = int(row_num)
    start_col = openpyxl.utils.column_index_from_string(col_letter)

    # Find max column in header row
    max_col = ws.max_column
    headers = []
    header_col_map = {}
    for col in range(start_col, max_col + 1):
        val = ws.cell(row=row_num, column=col).value
        if val is None:
            break
        headers.append(val)
        header_col_map[val] = col
    if not headers:
        print('No headers found at specified location.', file=sys.stderr)
        sys.exit(1)

    # If user specified headers to print, filter
    if args.print_headers:
        requested_headers = [h.strip() for h in args.print_headers.split(',') if h.strip()]
        missing = [h for h in requested_headers if h not in headers]
        if missing:
            print(f"Requested headers not found: {', '.join(missing)}", file=sys.stderr)
            sys.exit(1)
        headers_to_print = requested_headers
    else:
        headers_to_print = headers

    if args.only_headers:
        print("Headers:")
        for h in headers_to_print:
            print(h)
        sys.exit(0)

    # Collect data rows for selected headers
    # import already at top
    data_rows = []
    col_indices = [header_col_map[h] for h in headers_to_print]
    skip_col_idx = None
    found_formula = False
    if args.skip_if_empty:
        if args.skip_if_empty not in headers_to_print:
            print(f"Skip column header '{args.skip_if_empty}' not found in selected headers.", file=sys.stderr)
            sys.exit(1)
        skip_col_idx = headers_to_print.index(args.skip_if_empty)
    for row in ws.iter_rows(min_row=row_num+1, max_row=ws.max_row):
        data_row = []
        for i, col_idx in enumerate(col_indices):
            cell = row[col_idx-1]
            val = cell.value
            # Formula handling
            if isinstance(val, openpyxl.worksheet.formula.ArrayFormula):
                found_formula = True
                if args.formula_mode == 'print':
                    val = getattr(cell.value, 'formula', str(cell.value))
                else:
                    val = cell.internal_value if hasattr(cell, 'internal_value') else getattr(cell.value, 'formula', str(cell.value))
            data_row.append(val)
        if skip_col_idx is not None and (data_row[skip_col_idx] is None or str(data_row[skip_col_idx]).strip() == ""):
            continue
        data_rows.append(data_row)

    print_table(headers_to_print, data_rows)
    if found_formula:
        print("Note: One or more cells contain formulas. Formula values may not be calculated by openpyxl.")

if __name__ == '__main__':
    main()
