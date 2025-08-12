import argparse
import os
import requests
import openpyxl
import datetime

def get_jira_server_url():
    jira_server = os.environ.get("JIRA_SERVER_NAME")
    if not jira_server:
        raise EnvironmentError("JIRA_SERVER_NAME environment variable is not set.")
    return f"https://{jira_server}/rest/api/2"

def jira_request_bulk(jira_ids, token, field_ids):
    # Use Jira search API to fetch multiple issues at once
    jql = "key in ({})".format(",".join(jira_ids))
    url = f"{get_jira_server_url()}/search"
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "jql": jql,
        "fields": ",".join(field_ids)
    }
    response = requests.get(url, headers=headers, params=params)
    return response.json()["issues"] if response.ok else []

def jira_update_field(jira_id, field_id, new_value, token):
    url = f"{get_jira_server_url()}/issue/{jira_id}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    # Prepare the data dictionary according to field type
    # For custom fields, use appropriate format based on field_id
    if field_id == "project":
        data = {"fields": {"project": {"key": "ROSTER", "name": new_value}}}
    elif field_id == "issuetype":
        data = {"fields": {"issuetype": {"name": new_value}}}
    elif field_id in [
        "customfield_21611",
        "customfield_21609",
        "customfield_21607",
        "customfield_21605",
        "customfield_21603",
        "customfield_21601"
    ]:
        data = {"fields": {field_id: {"name": new_value}}}
    elif field_id == "customfield_12671":
        data = {"fields": {field_id: {"value": new_value}}}
    else:
        data = {"fields": {field_id: new_value}}

    response = requests.put(url, headers=headers, json=data)
    if response.ok:
        print(f"Updated Jira issue: {response.text}")
        return response.ok
    else:
        print(f"Failed to update Jira issue: {response.text}")
        return None

def jira_request_for_field(token):
    url = f"{get_jira_server_url()}/field"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url, headers=headers)
    return response.json() if response.ok else None

def get_jira_resource(jira_url, token, cache):
    if jira_url in cache:
        return cache[jira_url]
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(jira_url, headers=headers, timeout=10)
    data = response.json() if response.ok else None
    cache[jira_url] = data
    return data

def prompt_for_token():
    token = os.environ.get("JIRA_ACC_TOKEN")
    if not token:
        token = input("Enter your Jira Personal Access Token: ")
    return token


def clean_value(value):
    """
    Cleans a given value by removing newline characters and leading/trailing whitespace.

    Args:
        value: The input value to clean. Can be any type.

    Returns:
        str: The cleaned string representation of the input value. Returns an empty string if the input is falsy.
    """ 
    if not value:
        return ""
    return str(value).replace('\n', '').strip()


def trim_datetime_suffix(value):
    """
    Trims ' 00:00:00' from a string if it matches the 'YYYY-MM-DD 00:00:00' format.

    Args:
        value (str): The input string.

    Returns:
        str: The trimmed string if applicable, else the original string.
    """
    if isinstance(value, str) and value.endswith(" 00:00:00"):
        parts = value.split(" ")
        if len(parts) == 2:
            date_part = parts[0]
            try:
                # Check if date_part is a valid date
                datetime.datetime.strptime(date_part, "%Y-%m-%d")
                return date_part
            except ValueError:
                pass
    return value


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("-f", "--file", required=True, help="Excel file path")
    parser.add_argument("-s", "--sheet", required=True, help="Sheet name")
    args = parser.parse_args()

    token = prompt_for_token()
    wb = openpyxl.load_workbook(args.file)
    ws = wb[args.sheet]

    # Cache field metadata
    fields_response = jira_request_for_field(token)
    display_name_to_id = {f["name"]: f["id"] for f in fields_response} if fields_response else {}

    # Collect all Jira IDs and relevant field display names
    jira_ids = []
    row_map = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        jira_id = clean_value(row[1].value)
        if jira_id and str(jira_id).startswith("ROSTER-") and str(jira_id)[7:].isdigit():
            jira_ids.append(jira_id)
            row_map[jira_id] = row

    # Get all field display names from header row
    header_display_names = [ws.cell(row=2, column=col_idx+1).value for col_idx in range(2, 20)]
    field_ids = [display_name_to_id.get(name, "") for name in header_display_names if name]

    # Bulk fetch Jira issues
    issues = jira_request_bulk(jira_ids, token, field_ids)
    issue_map = {issue["key"]: issue for issue in issues}

    nested_cache = {}
    for jira_id in jira_ids:
        row = row_map[jira_id]
        print(f"Row {row[0].row} update is in processing...")
        jira_response = issue_map.get(jira_id)
        if jira_response:
            for col_idx, header_display_name in enumerate(header_display_names, start=2):
                field_id = display_name_to_id.get(header_display_name, "")
                excel_value = trim_datetime_suffix(clean_value(row[col_idx].value))
                
                jira_value = jira_response.get("fields", {}).get(field_id, "")
                if isinstance(jira_value, dict) and "self" in jira_value:
                    nested = get_jira_resource(jira_value["self"], token, nested_cache)
                    jira_value = clean_value(
                        nested.get("name") or
                        nested.get("displayName") or
                        nested.get("value") or ""
                    )
                if str(jira_value) != str(excel_value):
                    print(f"Updating Jira ID {jira_id} field '{header_display_name}' from '{jira_value}' to '{excel_value}'")
                    success = jira_update_field(jira_id, field_id, excel_value, token)
                    if success:
                        print("Update successful.")
                    else:
                        print("Update failed.")
        else:
            print(f"Failed to retrieve data for Jira ID: {jira_id}")

if __name__ == "__main__":
    main()
