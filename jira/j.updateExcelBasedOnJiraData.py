#!/usr/bin/env python3.12

import argparse
import os
import requests
import openpyxl
from datetime import datetime


def get_jira_issues(jira_ids, token):
    """Fetch Jira issues by their IDs.

    Args:
        jira_ids (str): Comma-separated string of Jira issue IDs.
        token (str): Jira API token.

    Returns:
        dict: A dictionary mapping Jira issue IDs to their details.
    """
    jira_server = os.environ.get("JIRA_SERVER_NAME", "")
    jql = "key in ({})".format(",".join(jira_ids))
    url = f"https://{jira_server}/rest/api/2/search"
    headers = {"Authorization": f"Bearer {token}"}
    params = {"jql": jql, "maxResults": len(jira_ids)}
    response = requests.get(url, headers=headers, params=params, timeout=10)
    if response.ok:
        issues = response.json().get("issues", [])
        return {issue["key"]: issue for issue in issues}
    return {}


def get_jira_resource(jira_url, token, cache):
    """ Fetch a Jira resource by its URL. """
    if jira_url in cache:
        return cache[jira_url]
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(jira_url, headers=headers, timeout=10)
    data = response.json() if response.ok else None
    cache[jira_url] = data
    return data


def get_jira_fields(token):
    """ Fetch Jira fields.  """
    jira_server = os.environ.get("JIRA_SERVER_NAME", "")
    url = f"https://{jira_server}/rest/api/2/field"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url, headers=headers, timeout=10)
    return response.json() if response.ok else None


def build_field_id_map(token):
    """ Build a mapping of Jira field names to their IDs. """
    fields = get_jira_fields(token)
    return {field.get("name"): field.get("id", "") for field in fields} if fields else {}


def get_field_value(fields_json, field_id):
    """ Get the value of a specific field from the fields JSON. """
    if not fields_json or field_id not in fields_json:
        return ""
    return fields_json[field_id]


def prompt_for_token():
    """ Prompt for Jira API token. """
    token = os.environ.get("JIRA_ACC_TOKEN")
    if not token:
        token = input("Enter your Jira Personal Access Token: ")
    return token


def clean_value(value):
    """
    Cleans a given value by removing newline characters and leading/trailing whitespace.

    Args:
        value: The input value to clean. Can be any type.

    Returns:
        str: The cleaned string representation of the input value. Returns an empty string if the input is falsy.
    """ 
    if not value:
        return ""
    return str(value).replace('\n', '').strip()


def update_excel_with_jira(excel_path, sheet_name, token, dry_run):
    """
    Updates an Excel workbook with data fetched from Jira issues.

    This function reads the specified Excel file and sheet, extracts Jira issue IDs from the sheet,
    fetches corresponding issue data from Jira using the provided token, and updates the sheet with
    relevant field values. Only rows with valid Jira IDs (format: 'ROSTER-<number>') are processed.
    The function supports updating columns C to T (Excel columns 3 to 20) based on header names in row 2.

    Args:
        excel_path (str): Path to the Excel workbook (.xlsx) to update.
        sheet_name (str): Name of the worksheet to update.
        token (str): Authentication token for accessing the Jira API.

    Returns:
        None

    Side Effects:
        - Modifies and saves the specified Excel file with updated Jira data.
        - Prints status and error messages to the console.
    """

    if not os.path.isfile(excel_path):
        print(f"File '{excel_path}' does not exist.")
        return

    if not excel_path.lower().endswith(".xlsx"):
        print("Only Excel Workbook (xlsx) format file is supported.")
        return

    wb = openpyxl.load_workbook(excel_path)
    print("Available sheets:", wb.sheetnames)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Please choose from: {wb.sheetnames}")
        return

    # Create a backup copy of the sheet before updating
    backup_time = datetime.now().strftime("%d-%b-%y.%H-%M")
    backup_sheet = wb.copy_worksheet(wb[sheet_name])
    backup_sheet.title = f"{sheet_name}_{backup_time}"
    wb.save(excel_path)
    print(f"Backup worksheet '{backup_sheet.title}' created inside workbook '{excel_path}'.")

    ws = wb[sheet_name]

    field_id_map = build_field_id_map(token)
    nested_cache = {}

    # Collect all Jira IDs to batch fetch
    jira_ids = []
    row_map = {}
    print("\n\nCollecting Jira IDs for batch fetching...")
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        print(f"Row {row[0].row} is processing to get the Jira ID...")
        jira_id = clean_value(row[1].value)
        if jira_id and str(jira_id).startswith("ROSTER-") and str(jira_id)[7:].isdigit():
            jira_ids.append(jira_id)
            row_map[jira_id] = row
        else:
            print(f"Row {row[0].row}: Invalid Jira ID format.")

    # Batch fetch issues
    print("\n\nFetching Jira issues from server...")
    issues = get_jira_issues(jira_ids, token)

    print("\n\nUpdating Excel sheet with Jira data...")
    for jira_id, row in row_map.items():
        print(f"Row {row[0].row} update is in processing...")
        issue = issues.get(jira_id)
        if issue:
            for col_idx in range(2, 20):  # C to T
                header_name = ws.cell(row=2, column=col_idx+1).value
                field_id = field_id_map.get(header_name, "")
                field_value = get_field_value(issue.get("fields", {}), field_id)
                if isinstance(field_value, dict) and "self" in field_value:
                    nested = get_jira_resource(field_value["self"], token, nested_cache)
                    field_value = (
                        nested.get("name") or
                        nested.get("displayName") or
                        nested.get("value") or ""
                    )
                if not dry_run:
                    ws.cell(row=row[0].row, column=col_idx+1).value = field_value
                else:
                    print(f"Would update row {row[0].row}, column {col_idx+1} ('{header_name}') with value: {field_value}")
        else:
            print(f"Could not retrieve data for Jira ID: {jira_id}")

    if not dry_run:
        wb.save(excel_path)
        print(f"\n\n{excel_path} Excel file updated successfully.")
    else:
        print(f"Dry run: No changes made to {excel_path}.")

def main():
    """ Update an Excel sheet with data from Jira. """
    parser = argparse.ArgumentParser(
        description="Update an Excel sheet with data from Jira based on Jira IDs in the sheet."
    )
    parser.add_argument("-f", "--file", required=True, help="Path to Excel file")
    parser.add_argument("-s", "--sheet", required=True, help="Sheet name to update")
    parser.add_argument("-d", "--dry-run", action="store_true", help="Print data only, do not create Jira issues")
    args = parser.parse_args()

    token = prompt_for_token()
    update_excel_with_jira(args.file, args.sheet, token, args.dry_run)


if __name__ == "__main__":
    main()
